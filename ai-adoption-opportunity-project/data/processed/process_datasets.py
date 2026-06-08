"""
process_datasets.py
===================
Cleans and processes source datasets for the
AI Adoption Opportunity project (Document Intelligence for Logistics).

Inputs  (place in data/raw/):
    LPI raw files (from https://databank.worldbank.org/data/download/LPI_CSV.zip):
        LPICSV.csv          LPI data — all countries, all indicators, all years
        LPICountry.csv      Country metadata — region, income group, etc.
        LPISeries.csv       Indicator metadata — definitions, licence, etc.

    Pre-curated project files:
        benchmarks.csv      AI vs manual performance benchmarks
        company_cases.csv   Operator case study results

Outputs (written to data/processed/ as Excel .xlsx):
    LPI_data.xlsx           Long-format data: one row per country × indicator × year
    LPI_countries.xlsx      Country metadata with clean column names
    LPI_indicators.xlsx     Indicator metadata with clean column names
    benchmarks.xlsx         Validated benchmarks — numeric columns typed correctly
    company_cases.xlsx      Validated case study results

    All outputs use .xlsx so Tableau reads column types (number, text, date)
    directly from Excel cell formatting — no manual type conversion needed.

Tableau connection:
    Connect each .xlsx file as a separate data source.
    Join LPI_data ↔ LPI_countries  on  country_code  = country_code
    Join LPI_data ↔ LPI_indicators on  indicator_code = indicator_code

Source:
    World Bank Logistics Performance Index
    License: Creative Commons Attribution 4.0 (CC BY 4.0)

Usage:
    python data/processed/process_datasets.py

Author:  AI Adoption Opportunity Project
Date:    June 2026

PowerShell (Windows):
    cd C:\\Users\\dilia\\OneDrive\\IronHack\\Projects\\Project5\\ai-adoption-opportunity-project
    python data\\processed\\process_datasets.py
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Paths ──────────────────────────────────────────────────────────────────
_SCRIPT_DIR   = Path(__file__).resolve().parent   # data/processed/
_PROJECT_ROOT = _SCRIPT_DIR.parent.parent          # <project_root>/

RAW_DIR       = _PROJECT_ROOT / "data" / "raw"
PROCESSED_DIR = _PROJECT_ROOT / "data" / "processed"
PROCESSED_DIR.mkdir(parents=True, exist_ok=True)

# ── Styling constants ──────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1D9E75")
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
BODY_FONT     = Font(size=11, name="Calibri")
CENTER        = Alignment(horizontal="center", vertical="center")
LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=False)


# ── Helpers ────────────────────────────────────────────────────────────────

def load_csv(filename: str) -> pd.DataFrame:
    """Load a CSV from data/raw/, trying utf-8 then latin-1 then cp1252."""
    path = RAW_DIR / filename
    if not path.exists():
        raise FileNotFoundError(
            f"\n[ERROR] File not found: {path}"
            f"\n        Place it in data/raw/ and re-run."
        )
    for encoding in ["utf-8", "latin-1", "cp1252"]:
        try:
            df = pd.read_csv(path, low_memory=False, encoding=encoding)
            if encoding != "utf-8":
                print(f"[INFO]  {filename} — loaded with {encoding} encoding")
            print(f"[LOAD]  {filename} — {df.shape[0]:,} rows × {df.shape[1]} columns")
            return df
        except UnicodeDecodeError:
            continue
    raise ValueError(f"[ERROR] Could not decode {filename}")


def report_missing(df: pd.DataFrame, label: str) -> None:
    """Print missing value counts for columns that have any nulls."""
    missing = df.isnull().sum()
    missing = missing[missing > 0]
    if missing.empty:
        print(f"[CHECK] {label} — no missing values")
    else:
        print(f"[CHECK] {label} — columns with missing values:")
        for col, n in missing.items():
            print(f"        {col}: {n:,} ({100 * n / len(df):.1f}%)")


def save_xlsx(
    df: pd.DataFrame,
    filename: str,
    numeric_cols: list[str] | None = None,
    integer_cols: list[str] | None = None,
    col_widths:   dict[str, int] | None = None,
    sheet_name:   str = "data",
) -> None:
    """
    Save a DataFrame to data/processed/ as a formatted .xlsx file.

    - Header row: green fill, white bold text
    - Numeric columns: typed as float cells with format '0.00'
    - Integer columns: typed as int cells with format '0'
    - Text columns: left-aligned
    - Auto column widths (overridable via col_widths)
    - Freeze panes on row 2
    - All cells use Calibri 11pt — same font Tableau reads cleanly
    """
    path = PROCESSED_DIR / filename
    numeric_cols = numeric_cols or []
    integer_cols = integer_cols or []

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        # ── Header row styling ─────────────────────────────────────────────
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill      = HEADER_FILL
            cell.font      = HEADER_FONT
            cell.alignment = LEFT

        # ── Body: enforce types and formatting per column ──────────────────
        for col_idx, col_name in enumerate(df.columns, 1):
            is_numeric  = col_name in numeric_cols
            is_integer  = col_name in integer_cols

            for row_idx in range(2, len(df) + 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font      = BODY_FONT
                cell.alignment = LEFT

                if is_numeric:
                    try:
                        cell.value         = float(cell.value) if cell.value is not None else None
                        cell.number_format = "0.00"
                    except (TypeError, ValueError):
                        pass
                elif is_integer:
                    try:
                        cell.value         = int(float(cell.value)) if cell.value is not None else None
                        cell.number_format = "0"
                    except (TypeError, ValueError):
                        pass

        # ── Column widths ──────────────────────────────────────────────────
        for col_idx, col_name in enumerate(df.columns, 1):
            letter = get_column_letter(col_idx)
            if col_widths and col_name in col_widths:
                ws.column_dimensions[letter].width = col_widths[col_name]
            else:
                # Auto-size: max of header length and a sample of values
                max_len = len(str(col_name))
                sample  = df[col_name].dropna().astype(str).head(50)
                if not sample.empty:
                    max_len = max(max_len, sample.str.len().max())
                ws.column_dimensions[letter].width = min(max_len + 4, 60)

        # ── Freeze header row ──────────────────────────────────────────────
        ws.freeze_panes = "A2"

    print(f"[SAVE]  {filename} — {len(df):,} rows × {len(df.columns)} cols → {path}")


# ═══════════════════════════════════════════════════════════════════════════
# FILE 1 — LPI_data.xlsx
# Source: LPICSV.csv
# Wide → long. One row per country × indicator × year.
# ═══════════════════════════════════════════════════════════════════════════

def process_lpi_data(filename: str = "LPICSV.csv") -> pd.DataFrame:
    """
    Clean LPICSV.csv and reshape wide → long.

    Output columns — LPI_data.xlsx:
        country_code        ISO3 — join key to LPI_countries.xlsx
        country_name        e.g. "Germany"
        indicator_code      World Bank code — join key to LPI_indicators.xlsx
        indicator_name      Full indicator name
        indicator_label     Snake_case label e.g. "customs_score"
        indicator_type      "score" (1–5) or "rank" (1 = best)
        year                Survey year (integer)
        value               Score or rank (number)
        dataset_source      "World Bank LPI"
    """
    print("\n── LPICSV.csv → LPI_data.xlsx ──────────────────────────────────────")
    df = load_csv(filename)

    required = {"Country Name", "Country Code", "Indicator Name", "Indicator Code"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"[ERROR] Missing columns in {filename}: {missing}")

    year_cols = sorted(
        [c for c in df.columns if str(c).strip().isdigit() and len(str(c).strip()) == 4],
        key=int
    )
    print(f"[INFO]  Survey years: {year_cols}")

    df_long = df.melt(
        id_vars    = ["Country Name", "Country Code", "Indicator Name", "Indicator Code"],
        value_vars = year_cols,
        var_name   = "year",
        value_name = "value",
    )
    df_long["year"]  = df_long["year"].astype(int)
    df_long["value"] = pd.to_numeric(df_long["value"], errors="coerce")

    before = len(df_long)
    df_long = df_long.dropna(subset=["value"])
    print(f"[DROP]  {before - len(df_long):,} null rows removed")

    df_long["indicator_type"] = df_long["Indicator Code"].apply(
        lambda c: "score" if str(c).endswith(".XQ") else "rank"
    )
    mask = df_long["indicator_type"] == "rank"
    df_long.loc[mask, "value"] = df_long.loc[mask, "value"].round(0)

    label_map = {
        "LP.LPI.OVRL.XQ": "lpi_overall_score", "LP.LPI.OVRL.RK": "lpi_overall_rank",
        "LP.LPI.CUST.XQ": "customs_score",      "LP.LPI.CUST.RK": "customs_rank",
        "LP.LPI.INFR.XQ": "infrastructure_score","LP.LPI.INFR.RK": "infrastructure_rank",
        "LP.LPI.ITRN.XQ": "intl_shipments_score","LP.LPI.ITRN.RK": "intl_shipments_rank",
        "LP.LPI.LOGS.XQ": "logistics_competence_score","LP.LPI.LOGS.RK": "logistics_competence_rank",
        "LP.LPI.TRAC.XQ": "tracking_tracing_score","LP.LPI.TRAC.RK": "tracking_tracing_rank",
        "LP.LPI.TIME.XQ": "timeliness_score",    "LP.LPI.TIME.RK": "timeliness_rank",
    }
    df_long["indicator_label"] = df_long["Indicator Code"].map(label_map).fillna(df_long["Indicator Code"])
    df_long["dataset_source"]  = "World Bank LPI"

    df_long = df_long.rename(columns={
        "Country Name":   "country_name",
        "Country Code":   "country_code",
        "Indicator Name": "indicator_name",
        "Indicator Code": "indicator_code",
    })

    df_long = df_long[[
        "country_code", "country_name",
        "indicator_code", "indicator_name", "indicator_label", "indicator_type",
        "year", "value", "dataset_source",
    ]].sort_values(["country_name", "indicator_label", "year"]).reset_index(drop=True)

    report_missing(df_long, "LPI_data")
    print(f"[INFO]  {df_long['country_code'].nunique()} countries · "
          f"{df_long['indicator_label'].nunique()} indicators · "
          f"{sorted(df_long['year'].unique())} years")

    save_xlsx(df_long, "LPI_data.xlsx",
              numeric_cols=["value"],
              integer_cols=["year"],
              col_widths={
                  "country_code": 14, "country_name": 24,
                  "indicator_code": 20, "indicator_name": 55,
                  "indicator_label": 28, "indicator_type": 14,
                  "year": 8, "value": 10, "dataset_source": 18,
              })
    return df_long


# ═══════════════════════════════════════════════════════════════════════════
# FILE 2 — LPI_countries.xlsx
# Source: LPICountry.csv
# ═══════════════════════════════════════════════════════════════════════════

def process_lpi_countries(filename: str = "LPICountry.csv") -> pd.DataFrame:
    """
    Clean LPICountry.csv.

    Output columns — LPI_countries.xlsx:
        country_code, country_short_name, country_long_name,
        alpha2_code, region, income_group, currency
    """
    print("\n── LPICountry.csv → LPI_countries.xlsx ─────────────────────────────")
    df = load_csv(filename)

    col_map = {
        "Country Code": "country_code",   "Short Name":   "country_short_name",
        "Long Name":    "country_long_name","2-alpha code": "alpha2_code",
        "Region":       "region",          "Income Group": "income_group",
        "Currency Unit":"currency",
    }
    available = {k: v for k, v in col_map.items() if k in df.columns}
    df_out = df[list(available.keys())].rename(columns=available).copy()

    before = len(df_out)
    df_out = df_out.drop_duplicates(subset=["country_code"])
    if len(df_out) < before:
        print(f"[DEDUP] Removed {before - len(df_out):,} duplicate rows")

    df_out = df_out.sort_values("country_short_name").reset_index(drop=True)

    report_missing(df_out, "LPI_countries")
    print(f"[INFO]  {len(df_out)} countries")

    save_xlsx(df_out, "LPI_countries.xlsx",
              col_widths={
                  "country_code": 14, "country_short_name": 28,
                  "country_long_name": 48, "alpha2_code": 12,
                  "region": 30, "income_group": 20, "currency": 20,
              })
    return df_out


# ═══════════════════════════════════════════════════════════════════════════
# FILE 3 — LPI_indicators.xlsx
# Source: LPISeries.csv
# ═══════════════════════════════════════════════════════════════════════════

def process_lpi_indicators(filename: str = "LPISeries.csv") -> pd.DataFrame:
    """
    Clean LPISeries.csv.

    Output columns — LPI_indicators.xlsx:
        indicator_code, indicator_label, indicator_name,
        long_definition, licence_type
    """
    print("\n── LPISeries.csv → LPI_indicators.xlsx ─────────────────────────────")
    df = load_csv(filename)

    col_map = {
        "Series Code":     "indicator_code",
        "Indicator Name":  "indicator_name",
        "Long definition": "long_definition",
        "License Type":    "licence_type",
    }
    available = {k: v for k, v in col_map.items() if k in df.columns}
    df_out = df[list(available.keys())].rename(columns=available).copy()

    label_map = {
        "LP.LPI.OVRL.XQ": "lpi_overall_score", "LP.LPI.OVRL.RK": "lpi_overall_rank",
        "LP.LPI.CUST.XQ": "customs_score",      "LP.LPI.CUST.RK": "customs_rank",
        "LP.LPI.INFR.XQ": "infrastructure_score","LP.LPI.INFR.RK": "infrastructure_rank",
        "LP.LPI.ITRN.XQ": "intl_shipments_score","LP.LPI.ITRN.RK": "intl_shipments_rank",
        "LP.LPI.LOGS.XQ": "logistics_competence_score","LP.LPI.LOGS.RK": "logistics_competence_rank",
        "LP.LPI.TRAC.XQ": "tracking_tracing_score","LP.LPI.TRAC.RK": "tracking_tracing_rank",
        "LP.LPI.TIME.XQ": "timeliness_score",    "LP.LPI.TIME.RK": "timeliness_rank",
    }
    df_out["indicator_label"] = df_out["indicator_code"].map(label_map).fillna(df_out["indicator_code"])

    col_order = ["indicator_code", "indicator_label", "indicator_name", "long_definition", "licence_type"]
    col_order = [c for c in col_order if c in df_out.columns]
    df_out = df_out[col_order].sort_values("indicator_label").reset_index(drop=True)

    report_missing(df_out, "LPI_indicators")
    print(f"[INFO]  {len(df_out)} indicators")

    save_xlsx(df_out, "LPI_indicators.xlsx",
              col_widths={
                  "indicator_code": 20, "indicator_label": 28,
                  "indicator_name": 55, "long_definition": 80, "licence_type": 16,
              })
    return df_out


# ═══════════════════════════════════════════════════════════════════════════
# FILES 4 & 5 — benchmarks.xlsx and company_cases.xlsx
# Pre-curated files — validate, enforce types, save as Excel.
# ═══════════════════════════════════════════════════════════════════════════

def process_benchmarks(filename: str = "benchmarks.csv") -> pd.DataFrame:
    """
    Validate benchmarks.csv and save as benchmarks.xlsx.
    Enforces: benchmark_id and source_year as integers,
              kpi_value (renamed from value) as float.
    """
    print(f"\n── {filename} → benchmarks.xlsx ────────────────────────────────────")
    df = load_csv(filename)

    df["benchmark_id"] = pd.to_numeric(df["benchmark_id"], errors="coerce").astype("Int64")
    df["source_year"]  = pd.to_numeric(df["source_year"],  errors="coerce").astype("Int64")
    df["kpi_value"]    = pd.to_numeric(df["value"],        errors="coerce")
    df = df.drop(columns=["value"])

    col_order = ["benchmark_id", "category", "metric", "scenario",
                 "kpi_value", "unit", "source", "source_year", "notes"]
    df = df[col_order]

    before = len(df)
    df = df.drop_duplicates()
    if len(df) < before:
        print(f"[DEDUP] Removed {before - len(df):,} duplicate rows")

    report_missing(df, "benchmarks")

    save_xlsx(df, "benchmarks.xlsx",
              numeric_cols=["kpi_value"],
              integer_cols=["benchmark_id", "source_year"],
              col_widths={
                  "benchmark_id": 14, "category": 22, "metric": 45,
                  "scenario": 14, "kpi_value": 12, "unit": 14,
                  "source": 30, "source_year": 13, "notes": 55,
              })
    return df


def process_company_cases(filename: str = "company_cases.csv") -> pd.DataFrame:
    """
    Validate company_cases.csv and save as company_cases.xlsx.
    Enforces result_value as float where possible.
    """
    print(f"\n── {filename} → company_cases.xlsx ─────────────────────────────────")
    df = load_csv(filename)

    if "result_value" in df.columns:
        df["result_value"] = pd.to_numeric(df["result_value"], errors="coerce")

    before = len(df)
    df = df.drop_duplicates()
    if len(df) < before:
        print(f"[DEDUP] Removed {before - len(df):,} duplicate rows")

    report_missing(df, "company_cases")

    numeric_cols = ["result_value"] if "result_value" in df.columns else []
    save_xlsx(df, "company_cases.xlsx",
              numeric_cols=numeric_cols)
    return df


# ═══════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("  AI Adoption Opportunity — Dataset Processing")
    print("  Document Intelligence for Logistics")
    print("=" * 60)

    results = {}

    # ── LPI files ─────────────────────────────────────────────────────────
    for raw_file, processor in [
        ("LPICSV.csv",     process_lpi_data),
        ("LPICountry.csv", process_lpi_countries),
        ("LPISeries.csv",  process_lpi_indicators),
    ]:
        try:
            results[raw_file] = processor(raw_file)
        except FileNotFoundError as e:
            print(str(e))
        except ValueError as e:
            print(str(e))

    # ── Pre-curated files ─────────────────────────────────────────────────
    for raw_file, processor in [
        ("benchmarks.csv",    process_benchmarks),
        ("company_cases.csv", process_company_cases),
    ]:
        try:
            results[raw_file] = processor(raw_file)
        except FileNotFoundError as e:
            print(str(e))

    # ── Summary ───────────────────────────────────────────────────────────
    print("\n" + "=" * 60)
    print("  Processing complete")
    print("=" * 60)
    for fname, df in results.items():
        out = fname.replace(".csv", ".xlsx")
        print(f"  ✓  {out:35} {df.shape[0]:,} rows × {df.shape[1]} cols")

    print(f"\n  Files saved to: {PROCESSED_DIR.resolve()}")
    print("\n  Tableau connection guide:")
    print("  1. Connect LPI_data.xlsx        — primary LPI data source")
    print("     Join LPI_countries.xlsx       on country_code  = country_code")
    print("     Join LPI_indicators.xlsx      on indicator_code = indicator_code")
    print("  2. Connect benchmarks.xlsx      — Panels 1–5 (filter on scenario, metric)")
    print("     Numeric column: kpi_value     (replaces 'value' to avoid reserved word)")
    print("  3. Connect company_cases.xlsx   — Panel 4 proof section")
    print("     Filter: relevance_to_muller = 'Direct'")
    print("\n  Useful LPI filters in Tableau:")
    print("  indicator_label = 'customs_score'  → Panel 1 bar chart")
    print("  indicator_type  = 'score'           → 1–5 scale values only")
    print("  year            = 2023              → latest survey round")


if __name__ == "__main__":
    main()
